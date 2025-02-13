
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from .models import Sede, Usuario,Evaluacion, Anotacion, Curso, Matricula,Asignatura,DiaSemana,AsignacionProfesorSede,Calificacion,Horario,PagoMensualidad,RegistroAsistencia
from django.contrib.auth.views import LoginView
from django.contrib.auth import logout
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from .forms import UsuarioForm,EditarUsuarioForm, AnotacionForm ,EvaluacionForm, InformeAsistenciaForm,SedeForm,CalificacionFormSet, CertificadoForm,CursoForm,ParametrosInformeAlumnoForm,ParametrosInformeForm,CalificacionSeleccionForm,AsignaturaForm,DiaSemanaForm,AsistenciaSeleccionForm,RegistroAsistenciaFormSet,MatriculaForm,AsignacionForm,HorarioForm, HorarioFiltroForm,PagoMensualidadForm, PagoMensualidadFiltroForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView,DetailView,FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Avg, Count, Q
from django.utils import timezone
import os
from django.http import FileResponse, HttpResponse, Http404
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from urllib.parse import urlencode
from django.utils.timezone import now
from django.forms import modelformset_factory
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
from decimal import Decimal
from django.http import JsonResponse
from datetime import datetime, date, timedelta
from django.views import View
from collections import defaultdict
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from .menus import MENUS 
from django.core.paginator import Paginator
from functools import wraps
from django.core.exceptions import PermissionDenied
from urllib.parse import unquote,quote
from django.contrib.staticfiles import finders
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from django.http import HttpResponseRedirect
from django.views.decorators.http import require_http_methods


def home(request):
    """Vista de página principal antes de iniciar sesión"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect(reverse_lazy('login'))


# Vista personalizada para el login (verifica si el usuario ya está autenticado)  

class CustomLoginView(LoginView):
    template_name = "colegio/login.html"  # Ruta correcta para la plantilla

    def get_redirect_url(self):
        # Si el usuario ya está autenticado, redirige al dashboard
        if self.request.user.is_authenticated:
            return reverse_lazy("dashboard")
        
        # Si no está autenticado, usa la URL por defecto
        return super().get_redirect_url()



# Vista personalizada para el logout (cerrar sesión)
def custom_logout_view(request):
    logout(request)
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect('login')  # Redirige a la página de login después de cerrar sesión
#----------------------------------------------------------

@login_required
def dashboard(request):
    if request.user.rol == 'PROFESOR':
        return redirect('dashboard_profesor')
    
    # Obtener estadísticas generales para todos los roles
    total_usuarios = Usuario.objects.count()
    total_cursos = Curso.objects.count()
    total_matriculas = Matricula.objects.count()
    
    # Obtener el menú correspondiente al rol del usuario
    rol = request.user.rol
    menu = MENUS.get(rol, [])
    
    context = {
        'total_usuarios': total_usuarios,
        'total_cursos': total_cursos,
        'total_matriculas': total_matriculas,
        'menu': menu,
    }
    
    return render(request, "colegio/dashboard.html", context)

#-----------------------------------------------------------

def es_profesor(user):
    return user.rol == 'PROFESOR'

@login_required
@user_passes_test(es_profesor)
def dashboard_profesor(request):
    # Obtener fecha actual o la seleccionada
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    
    if mes and ano:
        fecha_actual = timezone.make_aware(datetime(int(ano), int(mes), 1))
    else:
        fecha_actual = timezone.now()

    
    # Obtener primer y último día del mes
    # Obtener primer y último día del mes
    _, ultimo_dia = calendar.monthrange(fecha_actual.year, fecha_actual.month)

    if timezone.is_aware(fecha_actual):
        primer_dia = fecha_actual.replace(day=1)
        ultimo_dia = fecha_actual.replace(day=ultimo_dia)
    else:
        primer_dia = timezone.make_aware(fecha_actual.replace(day=1))
        ultimo_dia = timezone.make_aware(fecha_actual.replace(day=ultimo_dia))
    
    # Obtener evaluaciones del profesor para el mes seleccionado

    evaluaciones = Evaluacion.objects.filter(
        profesor=request.user,
        fecha__range=(primer_dia, ultimo_dia + timedelta(days=1))  # Agregamos un día más al rango
    ).select_related('asignatura', 'asignatura__curso')
    # Crear calendario
    cal = calendar.monthcalendar(fecha_actual.year, fecha_actual.month)
    cal_lunes_a_viernes = [
        semana[:5] for semana in cal
    ]
    

    # Organizar evaluaciones por día
    evaluaciones_por_dia = {}
    for evaluacion in evaluaciones:
        dia = evaluacion.fecha.astimezone(timezone.get_current_timezone()).day
        if dia not in evaluaciones_por_dia:
            evaluaciones_por_dia[dia] = []
        evaluaciones_por_dia[dia].append(evaluacion)
    
    # Estadísticas del profesor
    total_evaluaciones = Evaluacion.objects.filter(profesor=request.user).count()
    evaluaciones_mes = evaluaciones.count()
    proximas_evaluaciones = Evaluacion.objects.filter(
        profesor=request.user,
        fecha__gte=timezone.now()
    ).order_by('fecha')[:5]
    
    context = {
        'calendario': cal_lunes_a_viernes,
        'evaluaciones_por_dia': evaluaciones_por_dia,
        'mes_actual': fecha_actual,
        'mes_anterior': (fecha_actual.replace(day=1) - timedelta(days=1)),
        'mes_siguiente': (fecha_actual.replace(day=28) + timedelta(days=4)).replace(day=1),
        'total_evaluaciones': total_evaluaciones,
        'evaluaciones_mes': evaluaciones_mes,
        'proximas_evaluaciones': proximas_evaluaciones,
        'menu': MENUS.get('PROFESOR', [])  # Usar el menú importado
    }
    
    return render(request, 'colegio/dashboard_profesor.html', context)
#----------------------------------------------------------

#funciones de usuarios

def lista_usuarios(request):
    """Vista para listar usuarios"""
    usuarios = Usuario.objects.all().order_by("rol")  # Obtén todos los usuarios
    return render(request, 'colegio/lista_usuarios.html', {'usuarios': usuarios})


@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado correctamente.")
            return redirect('lista_usuarios')
        else:
            messages.error(request, "Corrige los errores en el formulario.")
    else:
        form = UsuarioForm()

    return render(request, 'colegio/crear_usuario.html', {'form': form})


@login_required
def detalle_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    return render(request, 'detalle_usuario.html', {'usuario': usuario})


@login_required
def editar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect('lista_usuarios')
        else:
            messages.error(request, "Corrige los errores en el formulario.")
    else:
        form = EditarUsuarioForm(instance=usuario)

    return render(request, 'colegio/editar_usuario.html', {'form': form})


@login_required
def eliminar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, "Usuario  eliminado exitosamente.")
        return redirect('lista_usuarios')       
    return render(request, 'confirmar_eliminar.html', {'usuario': usuario})


#funciones de Sedes

@login_required
def lista_sedes(request):
    sedes = Sede.objects.all()
    return render(request, 'colegio/lista_sedes.html', {'sedes': sedes})

@login_required
def crear_sede(request):
    if request.method == 'POST':
        form = SedeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_sedes')
    else:
        form = SedeForm()
    return render(request, 'colegio/formulario_sede.html', {'form': form})

@login_required
def editar_sede(request, pk):
    sede = get_object_or_404(Sede, pk=pk)
    if request.method == 'POST':
        form = SedeForm(request.POST, instance=sede)
        if form.is_valid():
            form.save()
            return redirect('lista_sedes')
    else:
        form = SedeForm(instance=sede)
    return render(request, 'colegio/formulario_sede.html', {'form': form})

@login_required
def eliminar_sede(request, pk):
    sede = get_object_or_404(Sede, pk=pk)
    if request.method == 'POST':
        sede.delete()
        return redirect('lista_sedes')
    return render(request, 'colegio/confirmar_eliminar_sede.html', {'sede': sede})


#Vista Cursos

@login_required
def lista_cursos(request):
    cursos = Curso.objects.all()
    return render(request, 'colegio/lista_cursos.html', {'cursos': cursos})

@login_required
def detalle_curso(request, curso_id):
    curso = get_object_or_404(Curso, pk=curso_id)
    return render(request, 'colegio/detalle_curso.html', {'curso': curso})

@login_required
def crear_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso creado exitosamente.')
            return redirect('lista_cursos')
    else:
        form = CursoForm()
    return render(request, 'colegio/crear_curso.html', {'form': form})

@login_required
def editar_curso(request, curso_id):
    curso = get_object_or_404(Curso, pk=curso_id)
    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso actualizado exitosamente.')
            return redirect('lista_cursos')
    else:
        form = CursoForm(instance=curso)
    return render(request, 'colegio/editar_curso.html', {'form': form, 'curso': curso})

@login_required
def eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, pk=curso_id)
    if request.method == 'POST':
        curso.delete()
        messages.success(request, 'Curso eliminado exitosamente.')
        return redirect('lista_cursos')
    return render(request, 'colegio/eliminar_curso.html', {'curso': curso})


#vistas asignaturas.

@login_required
def lista_asignaturas(request):
    query = request.GET.get('q', '')
    asignaturas = Asignatura.objects.all().order_by("curso__id", "nombre")

    if query:
        asignaturas = asignaturas.filter(
            Q(nombre__icontains=query) |
            Q(curso__nombre__icontains=query) |
            Q(profesor__first_name__icontains=query) |  # Buscar en el nombre del profesor
            Q(profesor__last_name__icontains=query)     # Buscar en el apellido del profesor
        )

    paginator = Paginator(asignaturas, 13)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    agrupadas_por_curso = {}
    for asignatura in page_obj:
        curso_nombre = asignatura.curso.nombre if asignatura.curso else "Sin curso"
        if curso_nombre not in agrupadas_por_curso:
            agrupadas_por_curso[curso_nombre] = []
        agrupadas_por_curso[curso_nombre].append(asignatura)

    return render(request, 'colegio/lista_asignaturas.html', {
        'page_obj': page_obj,
        'agrupadas_por_curso': agrupadas_por_curso,
        'query': query,
    })
    
@login_required
def detalle_asignatura(request, pk):
    asignatura = get_object_or_404(Asignatura, pk=pk)
    return render(request, 'colegio/detalle_asignatura.html', {
        'asignatura': asignatura
    })


@login_required
def crear_asignatura(request):
    if request.method == 'POST':
        form = AsignaturaForm(request.POST)
        if form.is_valid():
            form.save()  # Simplemente guarda sin asignar
            messages.success(request, 'Asignatura creada exitosamente.')
            return redirect('lista_asignaturas')
    else:
        form = AsignaturaForm()
    return render(request, 'colegio/crear_asignatura.html', {
        'form': form
    })

@login_required
def editar_asignatura(request, pk):
    asignatura = get_object_or_404(Asignatura, pk=pk)
    if request.method == 'POST':
        form = AsignaturaForm(request.POST, instance=asignatura)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignatura actualizada exitosamente.')
            return redirect('lista_asignaturas')
    else:
        form = AsignaturaForm(instance=asignatura)
    return render(request, 'colegio/editar_asignatura.html', {
        'form': form,
        'asignatura': asignatura
    })

@login_required
def eliminar_asignatura(request, pk):
    asignatura = get_object_or_404(Asignatura, pk=pk)
    if request.method == 'POST':
        asignatura.delete()
        messages.success(request, 'Asignatura eliminada exitosamente.')
        return redirect('lista_asignaturas')
    return render(request, 'colegio/eliminar_asignatura.html', {
        'asignatura': asignatura
    })
    
#dias de la semana

class DiaSemanaListView(ListView):
    model = DiaSemana
    template_name = 'colegio/lista_dias.html'
    context_object_name = 'dias'

class DiaSemanaCreateView(CreateView):
    model = DiaSemana
    form_class = DiaSemanaForm
    template_name = 'colegio/crear_dia.html'
    success_url = reverse_lazy('lista_dias')

class DiaSemanaUpdateView(UpdateView):
    model = DiaSemana
    form_class = DiaSemanaForm
    template_name = 'colegio/editar_dia.html'
    success_url = reverse_lazy('lista_dias')

class DiaSemanaDeleteView(DeleteView):
    model = DiaSemana
    template_name = 'colegio/confirmar_eliminar_dia.html'
    success_url = reverse_lazy('lista_dias')    
    

#vistas Matriculas


class MatriculaCreateView(LoginRequiredMixin, CreateView):
    model = Matricula
    form_class = MatriculaForm
    template_name = 'colegio/crear_matricula.html'
    success_url = reverse_lazy('lista_matriculas')

    def form_valid(self, form):
        # Asigna el usuario autenticado al campo usuario_creacion
        form.instance.usuario_creacion = self.request.user
        
        try:
            # Intenta guardar el formulario
            response = super().form_valid(form)
            messages.success(self.request, 'Matrícula creada exitosamente.')
            return response
        except ValidationError as e:
            # Maneja cualquier error de validación
            form.add_error(None, e)
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Añade un mensaje de error general
        messages.error(self.request, 'Hubo un error al crear la matrícula. Por favor, revise los campos.')
        return super().form_invalid(form)

class MatriculaUpdateView(LoginRequiredMixin, UpdateView):
    model = Matricula
    form_class = MatriculaForm
    template_name = 'colegio/editar_matricula.html'
    success_url = reverse_lazy('lista_matriculas')

    def form_valid(self, form):
        # Asigna el usuario autenticado al campo usuario_modificacion
        form.instance.usuario_modificacion = self.request.user
        
        try:
            # Intenta guardar el formulario
            response = super().form_valid(form)
            messages.success(self.request, 'Matrícula actualizada exitosamente.')
            return response
        except ValidationError as e:
            # Maneja cualquier error de validación
            form.add_error(None, e)
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Añade un mensaje de error general
        messages.error(self.request, 'Hubo un error al actualizar la matrícula. Por favor, revise los campos.')
        return super().form_invalid(form)

class MatriculaListView(LoginRequiredMixin, ListView):
    model = Matricula
    template_name = 'colegio/lista_matriculas.html'
    context_object_name = 'matriculas'
    
    def get_queryset(self):
        # Puedes añadir filtros si es necesario
        return Matricula.objects.all().order_by('-año', 'alumno')


class MatriculaDetailView(LoginRequiredMixin, DetailView):
    model = Matricula
    template_name = 'colegio/detalle_matricula.html'
    context_object_name = 'matricula'
    

class MatriculaDeleteView(LoginRequiredMixin, DeleteView):
    model = Matricula
    template_name = 'colegio/confirmar_eliminar_matricula.html'
    success_url = reverse_lazy('lista_matriculas')
    
#vistas asigana profesor a sede



class AsignacionListView(ListView):
    model = AsignacionProfesorSede
    template_name = 'colegio/asignacion_list.html'
    context_object_name = 'asignaciones'

class AsignacionCreateView(CreateView):
    model = AsignacionProfesorSede
    form_class = AsignacionForm
    template_name = 'colegio/crea_asignacionprofe.html'
    success_url = reverse_lazy('asignacion_list') #Redirige a la lista después de crear

class AsignacionUpdateView(UpdateView):
    model = AsignacionProfesorSede
    form_class = AsignacionForm
    template_name = 'colegio/crea_asignacionprofe.html'
    success_url = reverse_lazy('asignacion_list')#Redirige a la lista después de actualizar

class AsignacionDeleteView(DeleteView):
    model = AsignacionProfesorSede
    template_name = 'colegio/elimina_asignacionprofe.html'
    success_url = reverse_lazy('asignacion_list')#Redirige a la lista después de eliminar 
    

#calificaciones
@login_required
def lista_calificaciones(request):
    # Primero obtenemos las asignaturas del profesor
    asignaturas_profesor = Asignatura.objects.filter(profesor=request.user)
    
    # Luego obtenemos las calificaciones solo de esas asignaturas
    calificaciones = Calificacion.objects.filter(
        asignatura__in=asignaturas_profesor
    ).select_related(
        'matricula__alumno', 
        'asignatura',
        'matricula__curso'
    ).order_by(
        'asignatura__nombre',
        'matricula__curso',
        'matricula__alumno__last_name',
        'matricula__alumno__first_name',
        '-semestre',
        'tipo'
    )
    
    # Agrupamos las calificaciones por asignatura y curso para mejor organización
    calificaciones_agrupadas = {}
    for calificacion in calificaciones:
        key = (calificacion.asignatura.nombre, calificacion.matricula.curso)
        if key not in calificaciones_agrupadas:
            calificaciones_agrupadas[key] = []
        calificaciones_agrupadas[key].append(calificacion)
    
    return render(request, 'colegio/lista_calificaciones.html', {
        'calificaciones_agrupadas': calificaciones_agrupadas
    })
    

@login_required
def eliminar_calificacion(request, calificacion_id):
    calificacion = get_object_or_404(Calificacion, id=calificacion_id, profesor=request.user)
    
    if request.method == 'POST':
        calificacion.delete()
        return redirect('lista_calificaciones')
    
    return render(request, 'colegio/eliminar_calificacion.html', {
        'calificacion': calificacion
    })



@login_required
def horario_lista(request):
    # Consulta base de horarios
    horarios = Horario.objects.all()
    
    # Formulario de filtros
    form_filtro = HorarioFiltroForm(request.GET)
    
    # Aplicar filtros si son válidos
    if form_filtro.is_valid():
        curso = form_filtro.cleaned_data.get('curso')
        asignatura = form_filtro.cleaned_data.get('asignatura')
        profesor = form_filtro.cleaned_data.get('profesor')
        
        # Filtrar por cada campo si está presente
        if curso:
            horarios = horarios.filter(curso=curso)
        if asignatura:
            horarios = horarios.filter(asignatura=asignatura)
        if profesor:
            horarios = horarios.filter(asignacion_profesor_sede__usuario=profesor)
    
    # Incluir contexto de filtros y horarios
    context = {
        'horarios': horarios,
        'form_filtro': form_filtro
    }
    return render(request, 'colegio/horario_lista.html', context)

@login_required
def horario_crear(request):
    if request.method == 'POST':
        form = HorarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario creado exitosamente')
            return redirect('horario_lista')
    else:
        form = HorarioForm()
    
    return render(request, 'colegio/horario_form.html', {'form': form})

@login_required
def horario_editar(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    
    if request.method == 'POST':
        form = HorarioForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario actualizado exitosamente')
            return redirect('horario_lista')
    else:
        form = HorarioForm(instance=horario)
    
    return render(request, 'colegio/horario_form.html', {'form': form})

@login_required
def horario_eliminar(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    
    if request.method == 'POST':
        horario.delete()
        messages.success(request, 'Horario eliminado exitosamente')
        return redirect('horario_lista')
    
    return render(request, 'colegio/horario_confirmar_eliminar.html', {'horario': horario})



@login_required
def lista_pagos_mensualidad(request):
    # Obtiene todos los pagos con las relaciones necesarias
    pagos = PagoMensualidad.objects.all().select_related('matricula__alumno')

    # Inicializa el formulario con los datos de la solicitud GET
    form_filtro = PagoMensualidadFiltroForm(request.GET)

    # Verifica si el formulario es válido
    if form_filtro.is_valid():
        # Obtiene los datos filtrados
        alumno = form_filtro.cleaned_data.get('alumno')
        año = form_filtro.cleaned_data.get('año')
        mes = form_filtro.cleaned_data.get('mes')
        estado = form_filtro.cleaned_data.get('estado')

        # Aplica los filtros al queryset
        if alumno:
            pagos = pagos.filter(matricula=alumno)
        if año:
            pagos = pagos.filter(año=año)
        if mes:
            pagos = pagos.filter(mes=mes)
        if estado:
            pagos = pagos.filter(estado=estado)

    # Contexto para la plantilla
    context = {
        'pagos': pagos,
        'form_filtro': form_filtro
    }
    return render(request, 'colegio/lista_pagos.html', context)


@login_required
def crear_pago_mensualidad(request):
    if request.method == 'POST':
        form = PagoMensualidadForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago de mensualidad registrado exitosamente.')
            return redirect('lista_pagos_mensualidad')
        else:
            messages.success(request, 'ya se encuntra realizado un pago en este mes y año.')
    else:
        form = PagoMensualidadForm()
    
    return render(request, 'colegio/form_pago.html', {'form': form})

@login_required
def editar_pago_mensualidad(request, pk):
    pago = get_object_or_404(PagoMensualidad, pk=pk)
    
    if request.method == 'POST':
        form = PagoMensualidadForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago de mensualidad actualizado exitosamente.')
            return redirect('lista_pagos_mensualidad')
    else:
        form = PagoMensualidadForm(instance=pago)
    
    return render(request, 'colegio/form_pago.html', {'form': form})

@login_required
def eliminar_pago_mensualidad(request, pk):
    pago = get_object_or_404(PagoMensualidad, pk=pk)
    
    if request.method == 'POST':
        pago.delete()
        messages.success(request, 'Pago de mensualidad eliminado exitosamente.')
        return redirect('lista_pagos_mensualidad')
    
    return render(request, 'colegio/eliminar_pago.html', {'pago': pago})

#vouche de pago
def generar_voucher_pdf(request, pago_id):
    pago = get_object_or_404(PagoMensualidad, id=pago_id)
    
    # Ruta para guardar temporalmente el PDF
    pdf_path = str(settings.MEDIA_ROOT / 'vouchers' / f'voucher_{pago.id}.pdf')
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
   
    # Crear el PDF con ReportLab
    c = canvas.Canvas(pdf_path, pagesize=letter)
    
    try:
        # Configuración inicial
        c.setTitle(f"Voucher de Pago - {pago.matricula.alumno.get_full_name()}")
        
        # Título
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, 750, "Pago de Mensualidad Colegio más que Vencedores")
        
        # Contenido
        c.setFont("Helvetica", 12)
        y_position = 700  # Posición inicial Y
        
        # Lista de elementos a escribir
        elementos = [
            f"Alumno: {pago.matricula.alumno.get_full_name()}",
            f"Mes: {pago.get_mes_display()}",
            f"Año: {pago.año}",
            f"Monto: ${pago.monto}",
            f"Fecha de Pago: {pago.fecha_pago or 'Pendiente'}",
            f"Estado: {pago.get_estado_display()}",
            f"Generado el: {now().strftime('%d/%m/%Y')}"
        ]
        
        # Escribir cada elemento
        for elemento in elementos:
            c.drawString(50, y_position, elemento)
            y_position -= 25  # Espaciado entre líneas
        
        # Finalizar el PDF
        c.showPage()
        c.save()
        
        print(f"PDF generado exitosamente en: {pdf_path}")
        
    except Exception as e:
        print(f"Error generando PDF: {str(e)}")
        raise
    
    # Resto de tu código para WhatsApp y render...
    mensaje = f"Hola, este es tu comprobante de pago de mensualidad:\n" \
              f"Colegio mas que vencedores:\n" \
              f"Alumno: {pago.matricula.alumno.get_full_name()}\n" \
              f"Mes: {pago.get_mes_display()} Año: {pago.año}\n" \
              f"Monto: ${pago.monto}\n" \
              f"Estado: {pago.get_estado_display()}\n" \
              f"Fecha de Pago: {pago.fecha_pago or 'Pendiente'}"
    whatsapp_url = f"https://wa.me/?{urlencode({'text': mensaje})}"
   
    return render(request, 'colegio/voucher_detalle.html', {
        'pago': pago,
        'pdf_url': f'/pdf/{pago.id}/',
        'whatsapp_url': whatsapp_url,
    })
    
#ASITENCIA

class ListarAsistenciaView(LoginRequiredMixin,ListView):
    model = RegistroAsistencia
    template_name = 'colegio/listar_asistencia.html'
    context_object_name = 'asistencias'
    paginate_by = 50

    def get_queryset(self):
        queryset = RegistroAsistencia.objects.select_related(
            'matricula__curso',
            'matricula__alumno',
            'asignatura'
        )

        buscar = self.request.GET.get('buscar', '')
        if buscar:
            queryset = queryset.filter(
                Q(matricula__alumno__first_name__icontains=buscar) |
                Q(matricula__alumno__last_name__icontains=buscar) |
                Q(matricula__alumno__rut__icontains=buscar) |
                Q(asignatura__nombre__icontains=buscar) |
                Q(estado__icontains=buscar) |
                Q(fecha_hora__icontains=buscar)
            )

        return queryset.order_by(
            'matricula__curso__nombre',
            '-fecha_hora',
            'matricula__alumno__last_name'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['buscar'] = self.request.GET.get('buscar', '')
        
        # Agrupar asistencias por curso
        asistencias_por_curso = {}
        for asistencia in context['asistencias']:
            curso_nombre = asistencia.matricula.curso.nombre
            if curso_nombre not in asistencias_por_curso:
                asistencias_por_curso[curso_nombre] = []
            asistencias_por_curso[curso_nombre].append(asistencia)
        
        context['asistencias_por_curso'] = asistencias_por_curso
        return context

class EliminarAsistenciaView(DeleteView):
    model = RegistroAsistencia
    success_url = reverse_lazy('listar_asistencia')
    template_name = 'colegio/confirmar_eliminar_asistencia.html'

@login_required
def seleccionar_curso(request):
    if request.method == 'POST':
        form = AsistenciaSeleccionForm(request.POST, usuario=request.user)
        if form.is_valid():
            request.session['asignatura_id'] = form.cleaned_data['asignatura'].id
            return redirect('tomar_asistencia')
    else:
        form = AsistenciaSeleccionForm(usuario=request.user)
    
    return render(request, 'colegio/seleccionar_curso.html', {'form': form})


@login_required
def tomar_asistencia(request):
    asignatura_id = request.session.get('asignatura_id')
    if not asignatura_id:
        return redirect('seleccionar_curso')
    
    asignatura = Asignatura.objects.get(id=asignatura_id)
    matriculas = Matricula.objects.filter(curso=asignatura.curso)
    
    if request.method == 'POST':
        formset = RegistroAsistenciaFormSet(request.POST)
        if formset.is_valid():
            for form in formset:
                if form.is_valid():
                    registro = form.save(commit=False)
                    registro.asignatura = asignatura
                    registro.fecha_hora = timezone.now()
                    registro.save()
            messages.success(request, 'Asistencia registrada correctamente')
            return redirect('dashboard')
    else:
        initial = [{'matricula': matricula.id} for matricula in matriculas]
        formset = RegistroAsistenciaFormSet(initial=initial)
    
    return render(request, 'colegio/tomar_asistencia.html', {
        'formset': formset,
        'asignatura': asignatura
    })

#CALIFICACIONES
# Decorador que asegura que el usuario esté autenticado para acceder a esta vista
@login_required
def seleccionar_curso_calificacion(request):
    # Si la solicitud es POST (el usuario envió el formulario)
    if request.method == 'POST':
        # Creamos un formulario con los datos enviados y pasamos el usuario actual
        form = CalificacionSeleccionForm(request.POST, usuario=request.user)
        # Validamos el formulario
        if form.is_valid():
            # Guardamos en la sesión los datos seleccionados en el formulario
            request.session['calificacion_asignatura_id'] = form.cleaned_data['asignatura'].id
            request.session['calificacion_tipo'] = form.cleaned_data['tipo']
            request.session['calificacion_semestre'] = form.cleaned_data['semestre']
            request.session['calificacion_especificacion'] = form.cleaned_data.get('especificacion', None)

            # Redirigimos al usuario a la vista de ingreso de calificaciones
            return redirect('ingresar_calificaciones')
    else:
        # Si la solicitud es GET (el usuario accede por primera vez), creamos un formulario vacío
        form = CalificacionSeleccionForm(usuario=request.user)
    
    # Renderizamos la plantilla con el formulario para que el usuario lo complete
    return render(request, 'colegio/seleccionar_curso_calificacion.html', {'form': form})

#-----------------------------------------------------------------------
@login_required
def ingresar_calificaciones(request):
    # Recuperamos los datos seleccionados en la vista anterior
    asignatura_id = request.session.get('calificacion_asignatura_id')
    tipo = request.session.get('calificacion_tipo')
    semestre = request.session.get('calificacion_semestre')
    especificacion = request.session.get('calificacion_especificacion')
    
    if not all([asignatura_id, tipo, semestre]):
        return redirect('seleccionar_curso_calificacion')
   
    asignatura = Asignatura.objects.get(id=asignatura_id)
    matriculas = Matricula.objects.filter(
        curso=asignatura.curso,
        estado='ACTIVO'
    ).select_related('alumno').order_by('alumno__last_name', 'alumno__first_name')
    
    # Preparar calificaciones para cada alumno
    calificaciones_existentes = []
    for matricula in matriculas:
        calificacion, _ = Calificacion.objects.get_or_create(
            matricula=matricula,
            asignatura=asignatura,
            tipo=tipo,
            semestre=semestre,
            defaults={
                'profesor': request.user,
                'especificacion': especificacion,
                'nota': None
            }
        )
        calificaciones_existentes.append(calificacion)

    if request.method == 'POST':
        if 'guardar' in request.POST:
            formset = CalificacionFormSet(
                request.POST, 
                queryset=Calificacion.objects.filter(id__in=[c.id for c in calificaciones_existentes])
            )
            if formset.is_valid():
                instances = formset.save(commit=False)
                for instance in instances:
                    instance.profesor = request.user
                    instance.save()
                messages.success(request, 'Calificaciones guardadas correctamente')
                return redirect('dashboard')
        else:
            # Si se cancela, eliminamos las calificaciones que no tienen nota
            Calificacion.objects.filter(
                id__in=[c.id for c in calificaciones_existentes],
                nota__isnull=True
            ).delete()
            return redirect('dashboard')

    # Crear el formset con las calificaciones existentes
    formset = CalificacionFormSet(
        queryset=Calificacion.objects.filter(id__in=[c.id for c in calificaciones_existentes])
    )
    
    # Asociar formularios con nombres de estudiantes
    alumnos_forms = []
    for form, matricula in zip(formset.forms, matriculas):
        nombre_completo = f"{matricula.alumno.last_name} {matricula.alumno.first_name}"
        alumnos_forms.append((form, nombre_completo))

    return render(request, 'colegio/ingresar_calificaciones.html', {
        'formset': formset,
        'alumnos_forms': alumnos_forms,
        'asignatura': asignatura,
        'tipo': dict(Calificacion.TIPO_CHOICES)[tipo],
        'semestre': f"{semestre}° Semestre",
        'especificacion': especificacion
    })
#-------------------------------------------------------------------
# vistas de informes de notas

def roles_required(*roles):
   def decorator(view_func):
       @wraps(view_func)
       def _wrapped_view(request, *args, **kwargs):
           if request.user.rol in roles:
               return view_func(request, *args, **kwargs)
           messages.error(request, 'No tienes permisos para acceder a esta página.')
           return redirect('dashboard')
       return _wrapped_view
   return decorator

@login_required
@roles_required('PROFESOR', 'DIRECTOR')
def seleccionar_parametros_informe(request):
    if request.method == 'POST':
        form = ParametrosInformeForm(request.user, request.POST)
        if form.is_valid():
            asignatura_id = form.cleaned_data['asignatura'].id
            año = form.cleaned_data['año']
            semestre = form.cleaned_data['semestre']
            return redirect('generar_informe_notas', asignatura_id=asignatura_id, año=año, semestre=semestre)
    else:
        form = ParametrosInformeForm(request.user)
    
    return render(request, 'colegio/seleccionar_parametros.html', {
        'form': form
    })


@login_required
def generar_informe_notas(request, asignatura_id, año, semestre):
    # Verificar que la asignatura existe
    if request.user.rol == 'DIRECTOR':
        asignatura = get_object_or_404(Asignatura, id=asignatura_id)
    else:
        asignatura = get_object_or_404(Asignatura, id=asignatura_id, profesor=request.user)
    
    # Obtener calificaciones sin filtrar por profesor
    calificaciones = Calificacion.objects.filter(
        asignatura=asignatura,
        semestre=semestre
    ).order_by('matricula__alumno__last_name', 
               'matricula__alumno__first_name', 
               'tipo')
    
    # Resto del código se mantiene igual...    
    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        alignment=1,
        spaceAfter=30
    )
    
    # Título
    title = Paragraph(
        f"INFORME DE NOTAS - {asignatura.curso.nombre} - {asignatura.nombre}",
        title_style
    )
    elements.append(title)
    
    # Subtítulo con información adicional
    subtitle = Paragraph(
        f"Año: {año} - Semestre: {semestre}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 20))
    
    # Preparar datos para la tabla
    data = [['Alumno', 'Nota 1', 'Nota 2', 'Nota 3', 'Nota 4', 'Nota 5', 'Promedio']]
    
    # Agrupar calificaciones por alumno
    alumnos_data = {}
    for calif in calificaciones:
        alumno = calif.matricula.alumno
        if alumno not in alumnos_data:
            alumnos_data[alumno] = {'notas': {}, 'promedio': 0}
        alumnos_data[alumno]['notas'][calif.tipo] = calif.nota
    
    # Calcular promedios y llenar la tabla
    for alumno, datos in alumnos_data.items():
        notas = []
        suma = Decimal('0')
        count = 0
        for tipo in ['NOTA 1', 'NOTA 2', 'NOTA 3', 'NOTA 4', 'NOTA 5']:
            nota = datos['notas'].get(tipo, '')
            notas.append(str(nota) if nota else '-')
            if nota:
                suma += Decimal(str(nota))
                count += 1
        
        promedio = round(suma / count, 2) if count > 0 else '-'
        data.append([
            f"{alumno.last_name}, {alumno.first_name}",
            *notas,
            str(promedio)
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ]))
    
    elements.append(table)
    
    # Agregar firma del profesor
    elements.append(Spacer(1, 50))
    firma = Paragraph(f"""
    <para alignment="center">
    _______________________<br/>
    {request.user.get_full_name()}<br/>
    Profesor
    </para>
    """, styles['Normal'])
    elements.append(firma)
    
    # Generar PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"informe_notas_{asignatura.curso.nombre}_{asignatura.nombre}_S{semestre}_{año}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response



#informe de notas x alumnos


@login_required
def seleccionar_parametros_informe_alumno(request):
    if request.method == 'POST':
        form = ParametrosInformeAlumnoForm(request.POST)
        if form.is_valid():
            matricula = form.cleaned_data['alumno']
            alumno_id = matricula.alumno.id
            año = form.cleaned_data['año']
            semestre = form.cleaned_data['semestre']
            observaciones = form.cleaned_data['observaciones']

            # Codificar observaciones para la URL
            observaciones_encoded = quote(observaciones) if observaciones else ""

            # Construir la URL con parámetros
            url = reverse('generar_informe_notas_alumno', kwargs={
                'alumno_id': alumno_id,
                'año': año,
                'semestre': semestre
            })

            return redirect(f'{url}?observaciones={observaciones_encoded}')
    else:
        form = ParametrosInformeAlumnoForm()

    return render(request, 'colegio/seleccionar_parametros_alumno.html', {'form': form})

        

def load_alumnos_notas(request):
    curso_id = request.GET.get('curso')
    matriculas = Matricula.objects.filter(
        curso_id=curso_id,
        estado='ACTIVO'
    ).select_related('alumno')
    return JsonResponse(
        list(matriculas.values('id', 'alumno__first_name', 'alumno__last_name')), 
        safe=False
    )    
    
    
#---------------------


@login_required
def generar_informe_notas_alumno(request, alumno_id, año, semestre):
    # Obtener el alumno
    alumno = get_object_or_404(Usuario, id=alumno_id, rol='ALUMNO')
    
    # Obtener el curso del alumno a través de su matrícula
    matricula = Matricula.objects.filter(alumno=alumno, año=año).first()
    curso = matricula.curso if matricula else None
    logo_path = finders.find('img/icono.ico')  # Asegura que se busque dentro de STATICFILES_DIRS
    
    # Obtener todas las calificaciones del alumno
    calificaciones = Calificacion.objects.filter(
        matricula__alumno=alumno,
        semestre=semestre
    ).order_by('asignatura__nombre', 'tipo')
    
    # Calcular porcentaje de asistencia
    asistencias_totales = RegistroAsistencia.objects.filter(
        matricula=matricula,
        fecha_hora__year=año,
        fecha_hora__month__range=(1 if semestre == 1 else 7, 6 if semestre == 1 else 12)
    ).count()
    
    presentes = RegistroAsistencia.objects.filter(
        matricula=matricula,
        fecha_hora__year=año,
        fecha_hora__month__range=(1 if semestre == 1 else 7, 6 if semestre == 1 else 12),
        estado__in=['PRESENTE', 'JUSTIFICADO']
    ).count()
    
    porcentaje_asistencia = round((presentes / asistencias_totales * 100), 2) if asistencias_totales > 0 else 0
    
    # Decodificar las observaciones (pueden venir codificadas en la URL)
    observaciones = unquote(request.GET.get('observaciones', ''))
    
   
    # Crear el PDF con márgenes ajustados
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        leftMargin=0.75 * inch,   # Margen izquierdo
        rightMargin=0.75 * inch,  # Margen derecho
        topMargin=0.5 * inch,     # Margen superior reducido
        bottomMargin=0.75 * inch  # Margen inferior
    )
    elements = []   
    
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        alignment=1,
        spaceAfter=30,
        fontSize=18,
        textColor=colors.blue
    )
    
    # Agregar logo
    if logo_path:
        img = Image(logo_path)
        img.drawHeight = 50
        img.drawWidth = 50
        
        logo_table = Table([[img]], colWidths=50, rowHeights=50)
        logo_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('VALIGN', (0, 0), (0, 0), 'TOP'),
            ('LEFTPADDING', (0, 0), (0, 0), 0),
            ('TOPPADDING', (0, 0), (0, 0), 0),
        ]))
        elements.append(logo_table)
    
    # Título
    title = Paragraph(f"INFORME DE NOTAS - {semestre} <b>SEMESTRE</b>", title_style)
    elements.append(title)
    
    # Información del alumno
    profesor_jefe = curso.profesor_jefe.get_full_name() if curso and curso.profesor_jefe else "No asignado"
    nivel = curso.get_nivel_display() if curso else "No especificado"
    curso_nombre = curso.nombre if curso else "No especificado"
    curso_año = curso.año if curso else "No especificado"
    
    alumno_info_style = ParagraphStyle(
        'AlumnoInfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
    )
    alumno_info_style1 = ParagraphStyle(
        'AlumnoInfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        alignment=1,  # Centrar el texto
        fontName='Helvetica-Oblique'  # Usar fuente en cursiva
    )
    
    alumno_info = Paragraph(
        f"""
        <b>Alumno:</b> {alumno.get_full_name()}<br/>
        <b>Profesor Jefe:</b> {profesor_jefe}<br/>
        <b>Curso:</b> {curso_nombre}<br/>  
        <b>Año:</b> {curso_año}
        """,
        alumno_info_style
    )
    elements.append(alumno_info)
    
    # Espaciado
    elements.append(Paragraph("<br/>", styles['Normal']))
    elements.append(Spacer(1, 10))
    
    # Versículo bíblico
    alumno_info1 = Paragraph(
        f"""
        Y todo lo que hagáis, hacedlo de corazón, como para el Señor y no para los hombres;
        sabiendo que del Señor recibiréis la recompensa de la herencia, porque a Cristo el Señor servís.
        Colosenses 3:23-24
        """,
        alumno_info_style1
    )
    elements.append(alumno_info1)
      
    
    
    # Espaciado
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabla de notas
    data = [['Asignatura', 'Nota 1', 'Nota 2', 'Nota 3', 'Nota 4', 'Nota 5', 'Promedio']]
    
    # Agrupar calificaciones por asignatura
    asignaturas_data = {}
    for calif in calificaciones:
        asignatura = calif.asignatura
        if asignatura not in asignaturas_data:
            asignaturas_data[asignatura] = {'notas': {}, 'promedio': 0}
        asignaturas_data[asignatura]['notas'][calif.tipo] = calif.nota
    
    # Calcular promedios y llenar la tabla
    promedio_general = Decimal('0')
    asignaturas_con_notas = 0
    
    for asignatura, datos in asignaturas_data.items():
        notas = []
        suma = Decimal('0')
        count = 0
        for tipo in ['NOTA 1', 'NOTA 2', 'NOTA 3', 'NOTA 4', 'NOTA 5']:
            nota = datos['notas'].get(tipo, '')
            notas.append(str(nota) if nota else '-')
            if nota:
                suma += Decimal(str(nota))
                count += 1
        
        if count > 0:
            promedio = round(suma / count, 2)
            promedio_general += promedio
            asignaturas_con_notas += 1
        else:
            promedio = '-'
            
        data.append([
            asignatura.nombre,
            *notas,
            str(promedio) if promedio != '-' else '-'
        ])
    
    # Calcular promedio general
    if asignaturas_con_notas > 0:
        promedio_general = round(promedio_general / asignaturas_con_notas, 2)
        data.append(['Promedio General', '', '', '', '', '', str(promedio_general)])
    
    # Agregar fila de asistencia
    data.append(['Porcentaje de Asistencia', '', '', '', '', '', f"{porcentaje_asistencia}%"])
    
    # Crear tabla con estilos
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('BACKGROUND', (0, -2), (-1, -1), colors.HexColor('#D3DFEE')),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(table)
    
    # Observaciones
    if observaciones:
        elements.append(Spacer(1, 20))
        obs_style = ParagraphStyle(
            'Observaciones',
            parent=styles['Normal'],
            fontSize=11,
            leading=14,
            leftIndent=20,
            rightIndent=20
        )
        obs_title = Paragraph("<b>Observaciones:</b>", obs_style)
        obs_text = Paragraph(observaciones, obs_style)
        elements.append(obs_title)
        elements.append(obs_text)
    
    # Firma del profesor
    elements.append(Spacer(1, 30))
    firma = Paragraph(f"""
    <para alignment="center">
    _______________________<br/>
    {profesor_jefe}<br/>
    Profesor Jefe
    </para>
    """, styles['Normal'])
    elements.append(firma)
    
    # Pie de página
    footer = Paragraph(f"""
    <para alignment="center">
    <font size="9" color="gray">Escuela Mas Que Vencedores - Informe generado automáticamente</font>
    </para>
    """, styles['Normal'])
    elements.append(Spacer(1, 20))
    elements.append(footer)
    
    # Generar PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"informe_notas_{alumno.get_full_name()}_S{semestre}_{año}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
#-------------------  

#CERTIFICADO ALUMNO REGULAR
class CertificadoView(FormView):
    template_name = 'colegio/certificado_form.html'
    form_class = CertificadoForm
    success_url = reverse_lazy('certificado_form')

    def form_valid(self, form):
        matricula = form.cleaned_data['alumno']
        semestre = form.cleaned_data['semestre']
        
        # Configuración de la respuesta HTTP para PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="certificado.pdf"'

        # Crear el PDF
        p = canvas.Canvas(response)
        
        # Título del certificado
        p.setFont("Helvetica-Bold", 14)
        p.drawCentredString(300, 770, "CERTIFICADO DE ALUMNO REGULAR")
        #p.drawCentredString(300, 770, "COLEGIO MAS QUE VENCEDORES")

        # Cuerpo del certificado
        p.setFont("Helvetica", 12)
        p.drawString(50, 700, "     JESSICA GUERRA, Directora Colegio más que Vencedores Sede Colín. Región del Maule.")
        p.drawString(50, 680, f"Certifica que {matricula.alumno}, RUN N° {matricula.alumno.rut}, es alumno(a) regular")
        p.drawString(50, 660, f"de {matricula.curso.nombre} durante el período académico {matricula.año}.")
        
        p.drawString(50, 620, "Se extiende el presente certificado a petición del interesado(a) para: FINES QUE")
        p.drawString(50, 600, "ESTIME CONVENIENTES.")
        
       

        # Firma
        p.drawString(300, 450, "________________________")
        p.drawString(300, 430, "Jessica Guerra")
        p.drawString(300, 410, "Directora MQV")
         # Fecha
        p.drawString(50, 360, f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")

        # Validez
        p.drawString(50, 340, "Válido por 60 días.")
        p.drawString(50,50, "Departamento de Admisión y Registro Académico MQV.")
        p.drawString(50, 30, "Av. Colin S/N - Maule, Fono: 71-22464564 admision@mqv.cl")
        
        
        # Finalizar el PDF
        p.showPage()
        p.save()

        return response


#carga alumnos
def load_alumnos(request):
    curso_id = request.GET.get('curso')
    alumnos = Matricula.objects.filter(
        curso_id=curso_id,
        estado='ACTIVO'
    ).select_related('alumno')
    return JsonResponse(
        list(alumnos.values('id', 'alumno__first_name', 'alumno__last_name')), 
        safe=False
    )


# genera informe de asistencia 

def exportar_excel(context):
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    
    # Configurar el título
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Informe de Asistencia'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='left')
    
    # Configurar el subtítulo
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Asignatura: {context['asignatura']}, Año: {context['anio']}, Mes: {context['nombre_mes']}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='left')
    
    # Configurar encabezados
    headers = ['Alumno']
    fechas = context['fechas']
    for fecha in fechas:
        headers.append(str(fecha.day))
    headers.append('% Asistencia')
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        if col > 1 and col < len(headers):  # Solo para las fechas
            ws.column_dimensions[get_column_letter(col)].width = 4
    
    # Configurar ancho de columnas específicas
    ws.column_dimensions['A'].width = 30  # Columna de nombres
    ws.column_dimensions[get_column_letter(len(headers))].width = 12  # Columna de porcentaje
    
    # Escribir datos de asistencia
    row = 5
    for alumno, asistencias in context['alumnos_asistencia'].items():
        # Nombre del alumno
        ws.cell(row=row, column=1).value = f"{alumno.last_name}, {alumno.first_name}"
        
        # Estados de asistencia
        for col, fecha in enumerate(fechas, 2):
            cell = ws.cell(row=row, column=col)
            estado = asistencias[fecha]
            cell.value = estado
            cell.alignment = Alignment(horizontal='center')
            
            # Colorear según el estado
            if estado == 'P':
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde claro
            elif estado == 'A':
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Rojo claro
            elif estado == 'J':
                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # Amarillo claro
        
        # Porcentaje de asistencia
        porcentaje = context['estadisticas'][alumno]
        ws.cell(row=row, column=len(headers)).value = f"{porcentaje}%"
        ws.cell(row=row, column=len(headers)).alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=informe_asistencia.xlsx'
    
    # Guardar el libro de trabajo
    wb.save(response)
    
    return response
# genera informe de asistencia 
@login_required
def generar_informe(request):
    if request.method == 'POST':
        form = InformeAsistenciaForm(request.POST, usuario=request.user)
        if form.is_valid():
            try:
                asignatura = form.cleaned_data['asignatura']
                año = int(form.cleaned_data['año'])
                mes = int(form.cleaned_data['mes'])
                
                # Obtener fechas del mes (solo días de semana)
                primer_dia = date(año, mes, 1)
                ultimo_dia = date(año, mes, calendar.monthrange(año, mes)[1])
                fechas = [date(año, mes, dia) for dia in range(1, ultimo_dia.day + 1)
                         if date(año, mes, dia).weekday() < 5]
                
                # Obtener registros del mes
                registros = RegistroAsistencia.objects.filter(
                    asignatura=asignatura,
                    fecha_hora__date__range=(primer_dia, ultimo_dia)
                ).select_related('matricula__alumno').order_by(
                    'matricula__alumno__last_name',
                    'fecha_hora__date',
                    'fecha_hora__time'
                )
                
                # Crear diccionario de registros por alumno
                alumnos_dict = {}
                for registro in registros:
                    if registro.matricula.estado == 'ACTIVO':
                        alumno = registro.matricula.alumno
                        fecha = registro.fecha_hora.date()
                        
                        if alumno not in alumnos_dict:
                            alumnos_dict[alumno] = {fecha: [] for fecha in fechas}
                            
                        # Agregar el estado a la lista de estados del día
                        if fecha in alumnos_dict[alumno]:
                            alumnos_dict[alumno][fecha].append(registro.estado[0])
                
                # Consolidar múltiples estados por día
                for alumno in alumnos_dict:
                    for fecha in alumnos_dict[alumno]:
                        estados = alumnos_dict[alumno][fecha]
                        if estados:
                            if 'A' in estados:
                                alumnos_dict[alumno][fecha] = 'A'
                            elif 'P' in estados:
                                alumnos_dict[alumno][fecha] = 'P'
                            elif 'J' in estados:
                                alumnos_dict[alumno][fecha] = 'J'
                        else:
                            alumnos_dict[alumno][fecha] = '-'
                
                # Calcular estadísticas
                estadisticas = {}
                for alumno, asistencias in alumnos_dict.items():
                    presentes = sum(1 for estado in asistencias.values() if estado == 'P')
                    total_dias = len(fechas)
                    porcentaje = round((presentes / total_dias * 100), 1) if total_dias > 0 else 0
                    estadisticas[alumno] = porcentaje
                
                context = {
                    'form': form,
                    'asignatura': asignatura,
                    'anio': año,
                    'mes': mes,
                    'nombre_mes': calendar.month_name[mes],
                    'fechas': fechas,
                    'alumnos_asistencia': alumnos_dict,
                    'estadisticas': estadisticas,
                    'tiene_datos': bool(alumnos_dict)
                }
                
                if 'exportar' in request.POST:
                    return exportar_excel(context)
                    
                return render(request, 'colegio/informe_asistencia.html', context)
                
            except Exception as e:
                return render(request, 'colegio/generar_informeAsistencia.html', {
                    'form': form,
                    'error': f'Error al generar el informe: {str(e)}'
                })
    else:
        form = InformeAsistenciaForm(usuario=request.user)
    
    return render(request, 'colegio/generar_informeAsistencia.html', {'form': form})

#agendar evaluaciones


class EvaluacionCreateView(LoginRequiredMixin, CreateView):
    model = Evaluacion
    form_class = EvaluacionForm
    template_name = 'colegio/evaluacion_form.html'
    success_url = reverse_lazy('evaluacion_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['profesor'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.profesor = self.request.user
        try:
            messages.success(self.request, 'Evaluación agendada exitosamente')
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class EvaluacionListView(LoginRequiredMixin, ListView):
    model = Evaluacion
    template_name = 'colegio/evaluacion_list.html'
    context_object_name = 'evaluaciones'
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user
        
        # Filtro base de evaluaciones
        if user.rol == 'DIRECTOR':
            queryset = Evaluacion.objects.all()
        elif user.rol == 'PROFESOR':
            queryset = Evaluacion.objects.filter(profesor=user)
        else:
            return Evaluacion.objects.none()

        # Búsqueda y filtrado
        busqueda = self.request.GET.get('busqueda', '')
        if busqueda:
            try:
                # Intenta convertir la fecha al formato correcto
                fecha_convertida = datetime.strptime(busqueda.replace('/', '-'), '%d-%m-%Y').date()
                busqueda_fecha = fecha_convertida.strftime('%Y-%m-%d')
            except ValueError:
                busqueda_fecha = busqueda

            queryset = queryset.filter(
                Q(asignatura__nombre__icontains=busqueda) |
                Q(profesor__username__icontains=busqueda) |
                Q(profesor__first_name__icontains=busqueda) |
                Q(profesor__last_name__icontains=busqueda) |
                Q(fecha__date__icontains=busqueda_fecha) |
                Q(asignatura__curso__nombre__icontains=busqueda)
            ) 
        # Ordenar por fecha descendente
        return queryset.order_by('-fecha')


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Agrupar evaluaciones por curso
        evaluaciones_por_curso = {}
        for evaluacion in context['evaluaciones']:
            curso = evaluacion.asignatura.curso
            curso_nombre = curso.nombre if curso else 'Sin Curso'
            if curso_nombre not in evaluaciones_por_curso:
                evaluaciones_por_curso[curso_nombre] = []
            evaluaciones_por_curso[curso_nombre].append(evaluacion)
        
        context['evaluaciones_por_curso'] = evaluaciones_por_curso
        context['es_director'] = self.request.user.rol == 'DIRECTOR'
        
        return context

class EvaluacionUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Evaluacion
    form_class = EvaluacionForm
    template_name = 'colegio/evaluacion_form.html'
    success_url = reverse_lazy('evaluacion_list')

    def test_func(self):
        evaluacion = self.get_object()
        return self.request.user == evaluacion.profesor

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['profesor'] = self.request.user
        return kwargs

class EvaluacionDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Evaluacion
    template_name = 'colegio/evaluacion_confirm_delete.html'
    success_url = reverse_lazy('evaluacion_list')

    def test_func(self):
        evaluacion = self.get_object()
        return self.request.user == evaluacion.profesor

class EvaluacionDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Evaluacion
    template_name = 'colegio/evaluacion_detail.html'

    def test_func(self):
        evaluacion = self.get_object()
        return self.request.user == evaluacion.profesor
    
    

class TodasEvaluacionListView(LoginRequiredMixin, ListView):
    model = Evaluacion
    template_name = 'colegio/otrasevaluaciones.html'
    context_object_name = 'evaluaciones'
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user
        
        # Filtro base de evaluaciones
        if user.rol == 'DIRECTOR':
            queryset = Evaluacion.objects.all()
        elif user.rol == 'PROFESOR':
           queryset = Evaluacion.objects.all()
        else:
            return Evaluacion.objects.none()

        # Búsqueda y filtrado
        busqueda = self.request.GET.get('busqueda', '')
        if busqueda:
            try:
                # Intenta convertir la fecha al formato correcto
                fecha_convertida = datetime.strptime(busqueda.replace('/', '-'), '%d-%m-%Y').date()
                busqueda_fecha = fecha_convertida.strftime('%Y-%m-%d')
            except ValueError:
                busqueda_fecha = busqueda

            queryset = queryset.filter(
                Q(asignatura__nombre__icontains=busqueda) |
                Q(profesor__username__icontains=busqueda) |
                Q(profesor__first_name__icontains=busqueda) |
                Q(profesor__last_name__icontains=busqueda) |
                Q(fecha__date__icontains=busqueda_fecha) |
                Q(asignatura__curso__nombre__icontains=busqueda)
            ) 
        # Ordenar por fecha descendente
        return queryset.order_by('-fecha')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Agrupar evaluaciones por curso
        evaluaciones_por_curso = {}
        for evaluacion in context['evaluaciones']:
            curso = evaluacion.asignatura.curso
            curso_nombre = curso.nombre if curso else 'Sin Curso'
            if curso_nombre not in evaluaciones_por_curso:
                evaluaciones_por_curso[curso_nombre] = []
            evaluaciones_por_curso[curso_nombre].append(evaluacion)
        
        context['evaluaciones_por_curso'] = evaluaciones_por_curso
        
        
        return context
#ANOTACIONES



from django.core.paginator import Paginator
from django.db.models import Q

@login_required
def lista_anotaciones(request):
    query = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)

    if request.user.rol in ['PROFESOR', 'DIRECTOR', 'ADMIN']:
        anotaciones = Anotacion.objects.all()
    else:
        anotaciones = Anotacion.objects.filter(alumno=request.user)

    # Filtro por búsqueda
    if query:
        anotaciones = anotaciones.filter(
            Q(alumno__first_name__icontains=query) |
            Q(alumno__last_name__icontains=query) |
            Q(curso__nombre__icontains=query) |
            Q(nivel__icontains=query) |
            Q(descripcion__icontains=query)
        )

    # Ordenar por fecha
    anotaciones = anotaciones.order_by('-fecha_creacion')

    # Paginación
    paginator = Paginator(anotaciones, 10)  # 10 anotaciones por página
    anotaciones_paginadas = paginator.get_page(page_number)

    return render(request, 'lista_anotaciones.html', {
        'anotaciones': anotaciones_paginadas,
        'query': query
    })

   
@login_required
def detalle_anotacion(request, pk):
    anotacion = get_object_or_404(Anotacion, pk=pk)
    return render(request, 'detalle_anotacion.html', {'anotacion': anotacion})


class EditarAnotacionView(LoginRequiredMixin, UpdateView):
    model = Anotacion
    form_class = AnotacionForm
    template_name = 'editar_anotacion.html'
    success_url = reverse_lazy('lista_anotaciones')

    def form_valid(self, form):
        try:
            form.instance.usuario = self.request.user
            self.object = form.save()
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Anotación actualizada exitosamente',
                    'redirect_url': self.get_success_url()
                })
            return super().form_valid(form)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
            raise

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = [str(error) for error in error_list]
            
            return JsonResponse({
                'success': False,
                'message': 'Por favor corrija los errores en el formulario',
                'errors': errors
            }, status=400)
        return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        if self.request.method in ('POST', 'PUT'):
            kwargs['data'] = self.request.POST.copy()
        return kwargs


@login_required
def eliminar_anotacion(request, pk):
    if request.user.rol not in ['PROFESOR', 'DIRECTOR', 'ADMIN']:
        messages.error(request, 'No tiene permisos para eliminar anotaciones.')
        return redirect('lista_anotaciones')
        
    anotacion = get_object_or_404(Anotacion, pk=pk)
    if request.method == 'POST':
        anotacion.delete()
        messages.success(request, 'Anotación eliminada exitosamente.')
        return redirect('lista_anotaciones')
    return render(request, 'eliminar_Anotacion.html', {'anotacion': anotacion})


class CrearAnotacionView(LoginRequiredMixin, CreateView):
    model = Anotacion
    form_class = AnotacionForm
    template_name = 'crear_anotacion.html'
    success_url = reverse_lazy('lista_anotaciones')

    def form_valid(self, form):
        try:
            form.instance.usuario = self.request.user
            self.object = form.save()           
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Anotación guardada exitosamente',
                    'redirect_url': self.get_success_url()
                })
            return super().form_valid(form)
            
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
            raise

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = [str(error) for error in error_list]
            
            return JsonResponse({
                'success': False,
                'message': 'Por favor corrija los errores en el formulario',
                'errors': errors
            }, status=400)
        return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        if self.request.method in ('POST', 'PUT'):
            kwargs['data'] = self.request.POST.copy()
        return kwargs
    
def get_alumnos_curso(request):
    curso_id = request.GET.get('id_curso')
    if curso_id:
        try:
            alumnos = Usuario.objects.filter(
                rol='ALUMNO',
                matriculas__curso_id=curso_id,
                matriculas__estado='ACTIVO'  # Solo alumnos activos
            ).distinct().values('id', 'first_name', 'last_name')
            
            return JsonResponse({
                'success': True,
                'data': list(alumnos)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    return JsonResponse({
        'success': True,
        'data': []
    })